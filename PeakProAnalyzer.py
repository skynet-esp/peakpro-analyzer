# -*- coding: utf-8 -*-

import tkinter as tk
import sys
import os
import openpyxl
from itertools import zip_longest
from tkinter import filedialog, ttk, messagebox, simpledialog
import numpy as np
import json
import pickle
from Bio import SeqIO
from scipy.interpolate import interp1d
from scipy.signal import find_peaks
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from pathlib import Path


# --- Constantes y Configuración ---
CHANNEL_COLOR_MAP = {
    'DATA4':  'red',
    'DATA9':  'blue',
    'DATA10': 'green',
    'DATA11': 'black',
    'DATA1':  'blue',
    'DATA2':  'green',
    'DATA3':  'black',
}

# --- AÑADE ESTE DICCIONARIO AQUÍ ---
CHANNEL_DISPLAY_NAME_MAP = {
    'DATA9':  'Azul',
    'DATA10': 'Verde',
    'DATA11': 'Negro',
    'DATA4':  'Rojo (Marcador)',
    'DATA1':  'Azul',
    'DATA2':  'Verde',
    'DATA3':  'Negro',
}

KNOWN_LADDERS = {
    "GeneScan 500(-250) ROX": [35, 50, 75, 100, 139, 150, 160, 200, 250, 300, 340, 350, 400, 450, 490, 500],
    "BTO 550": [60, 80, 90, 100, 120, 140, 160, 180, 200, 220, 240, 250, 260, 280, 300, 320, 340, 360, 380, 400, 425, 450, 475, 500, 525, 550],
    "BTO 560": [73, 88, 123, 148, 173, 198, 223, 248, 273, 298, 324, 349, 373, 398, 423, 448, 470, 495, 520, 545, 555]
}

def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para desarrollo y para PyInstaller """
    try:
        # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- Clase SelectPeakDialog ---
class SelectPeakDialog(simpledialog.Dialog):
    def __init__(self, parent, title, available_sizes):
        self.available_sizes = available_sizes
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Selecciona el tamaño correcto para este pico:").pack(pady=5)
        self.result_var = tk.StringVar()
        self.combobox = ttk.Combobox(master, textvariable=self.result_var, values=self.available_sizes, state="readonly")
        if self.available_sizes:
            self.combobox.current(0)
        self.combobox.pack(pady=10, padx=10)
        return self.combobox 

    def apply(self):
        result_str = self.result_var.get()
        if result_str == "BORRAR ASIGNACIÓN":
            self.result = "DELETE"
        else:
            try:
                self.result = float(result_str)
            except (ValueError, TypeError):
                self.result = None

# --- Clase para la Portada (Splash Screen) con Botón de Entrada ---

class SplashScreen(tk.Toplevel):
    def __init__(self, parent, launch_callback, width=900, height=750):
        super().__init__(parent)
        self.launch_callback = launch_callback

        # Cargar la imagen de portada USANDO resource_path
        try:
            path_portada = resource_path("portada.png")
            self.bg_image = tk.PhotoImage(file=path_portada)
        except tk.TclError:
            self.bg_image = tk.PhotoImage(width=width, height=height)

        # Configuración de la ventana de la portada
        self.overrideredirect(True)
        
        # Centrar la ventana
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        self.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

        # Mostrar imagen y texto
        background_label = tk.Label(self, image=self.bg_image)
        background_label.place(x=0, y=0, relwidth=1, relheight=1)

        # --- Textos Personalizados ---
        text_frame = tk.Frame(self, bg='#222222')
        
        title_font = ("Segoe UI", 28, "bold")
        name_font = ("Segoe UI", 12, "italic")
        
        ttk.Label(text_frame, text="PeakPro Analyzer", font=title_font, foreground="#00BFFF", background='#222222').pack(pady=(10, 5))
        ttk.Label(text_frame, text="Jordi Martínez Serra @ Todos los derechos reservados", font=name_font, foreground="white", background='#222222').pack(pady=(0, 10))
        text_frame.pack(pady=60)

        # --- Contenedor para los botones ---
        button_frame = ttk.Frame(self)
        button_frame.pack(side=tk.BOTTOM, pady=40)

        # --- Botón de Salida ---
        # Le aplicamos el mismo estilo que al de Entrar para que se vean uniformes.
        exit_button = ttk.Button(button_frame, text="Salir", command=self.master.destroy, style="Enter.TButton")
        exit_button.pack(side=tk.LEFT, padx=30)
        
        # --- Botón de Entrada ---
        enter_button = ttk.Button(button_frame, text="Entrar", command=self.launch_callback, style="Enter.TButton")
        enter_button.pack(side=tk.RIGHT, padx=30)
        
# --- Asistente de Calibración ---
class CalibrationWizard(tk.Toplevel):
    def __init__(self, master, app, template=None):
        super().__init__(master)
        self.transient(master)
        self.grab_set()
        self.app = app
        self.current_file_index = 0
        self.calibrations = {}
        self.raw_data = None
        self.detected_peaks_indices = np.array([])
        self.manual_assignments = {}
        self.first_sample_template = template
        self.create_widgets()
        self.setup_for_current_file()
        self.fig.canvas.mpl_connect('button_press_event', self._on_plot_click)

    # En la clase CalibrationWizard, reemplaza esta función:

    # En la clase CalibrationWizard, reemplaza esta función:

    def create_widgets(self):
        self.title_var = tk.StringVar()
        self.title(self.title_var.get())
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.rowconfigure(0, weight=1); self.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1); main_frame.columnconfigure(0, weight=1)
        
        controls_frame = ttk.LabelFrame(main_frame, text="Parámetros de Detección", padding="10")
        controls_frame.grid(row=0, column=0, sticky="ew", pady=5)

        # Fila 1 de controles
        ttk.Label(controls_frame, text="Ignorar scans hasta:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.ignore_scans_var = tk.StringVar(value="1500")
        ttk.Entry(controls_frame, textvariable=self.ignore_scans_var, width=8).grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(controls_frame, text="Ancho Máximo (Picos):").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.width_var = tk.StringVar(value="100") 
        ttk.Entry(controls_frame, textvariable=self.width_var, width=8).grid(row=0, column=3, padx=5, pady=2)
        
        # --- NUEVO PARÁMETRO DE TOLERANCIA ---
        ttk.Label(controls_frame, text="Tolerancia Plantilla:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.template_tolerance_var = tk.StringVar(value="40") # Aumentamos el valor por defecto
        ttk.Entry(controls_frame, textvariable=self.template_tolerance_var, width=8).grid(row=0, column=5, padx=5, pady=2)

        # Fila 2 de controles
        ttk.Label(controls_frame, text="Altura Mínima:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.height_var = tk.StringVar(value="50")
        ttk.Entry(controls_frame, textvariable=self.height_var, width=8).grid(row=1, column=1, padx=5)

        ttk.Label(controls_frame, text="Prominencia Mínima:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.prominence_var = tk.StringVar(value="25")
        ttk.Entry(controls_frame, textvariable=self.prominence_var, width=8).grid(row=1, column=3, padx=5)
        
        ttk.Label(controls_frame, text="Distancia Mínima:").grid(row=1, column=4, padx=5, pady=5, sticky="w")
        self.distance_var = tk.StringVar(value="10")
        ttk.Entry(controls_frame, textvariable=self.distance_var, width=8).grid(row=1, column=5, padx=5)
        
        # Botón de refresco
        ttk.Button(controls_frame, text="Detectar / Refrescar Picos", command=lambda: self.detect_peaks(silent=False)).grid(row=2, column=0, columnspan=6, pady=10, sticky="ew")
        
        # El resto de la función se queda igual
        plot_frame = ttk.LabelFrame(main_frame, text="Gráfico Interactivo", padding="10"); plot_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        plot_frame.rowconfigure(0, weight=1); plot_frame.columnconfigure(0, weight=1)
        self.fig = plt.Figure(figsize=(10, 6)); self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame); self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2Tk(self.canvas, plot_frame, pack_toolbar=False); self.toolbar.update(); self.toolbar.pack(side=tk.TOP, fill=tk.X)
        action_frame = ttk.Frame(main_frame); action_frame.grid(row=2, column=0, pady=10, sticky="e")
        ttk.Button(action_frame, text="Limpiar Asignaciones", command=self.clear_assignments).pack(side=tk.LEFT, padx=5)
        self.next_button = ttk.Button(action_frame, command=self.save_and_next, style="Accent.TButton"); self.next_button.pack(side=tk.LEFT, padx=5)

    # En la clase CalibrationWizard, haz este pequeño cambio:

    def setup_for_current_file(self):
        self.clear_assignments(full_reset=True)
        full_path = self.app.fsa_files[self.current_file_index]
        filename_key = Path(full_path).name
        title = f"Calibrando Muestra {self.current_file_index + 1} de {len(self.app.fsa_files)}: {filename_key}"
        self.title(title)
        ladder_ch = self.app.ladder_channel_var.get()
        if ladder_ch not in self.app.loaded_data.get(filename_key, {}):
            messagebox.showerror("Error", f"El canal marcador '{ladder_ch}' no se encuentra en {filename_key}.", parent=self)
            self.raw_data = None
        else:
            self.raw_data = self.app.loaded_data[filename_key][ladder_ch]
        if self.current_file_index == len(self.app.fsa_files) - 1:
            self.next_button.config(text="Finalizar")
        else:
            self.next_button.config(text="Guardar y Siguiente")
            
        self.detect_peaks(silent=True) # Esto solo detecta los picos
        if self.first_sample_template:
            self._auto_assign_from_template(silent=True) # Y esto asigna la primera vez, sin mensaje
            
        self.redraw_plot()
    # En la clase CalibrationWizard, reemplaza esta función:


    def _auto_assign_from_template(self, silent=False):
        if not self.first_sample_template or self.raw_data is None or len(self.detected_peaks_indices) == 0:
            return
            
        try:
            tolerance = int(self.template_tolerance_var.get())
        except (ValueError, tk.TclError):
            tolerance = 40

        assigned_count = 0
        
        # 1. Preparamos las dos listas, ambas ordenadas
        # La plantilla, ordenada por tamaño de fragmento (bp)
        template_peaks_sorted = sorted(self.first_sample_template.items())
        # Los picos detectados, ordenados por posición (scan point)
        detected_peaks_sorted = sorted(list(self.detected_peaks_indices))

        # Índice para no volver a usar un pico detectado
        last_detected_idx = 0

        # 2. Iteramos por la plantilla en orden
        for bp_size, template_scan_point in template_peaks_sorted:
            
            # El espacio de búsqueda son solo los picos detectados que aún no hemos usado
            search_space = detected_peaks_sorted[last_detected_idx:]
            if len(search_space) == 0:
                break # No quedan más picos que asignar

            # Buscamos la distancia al pico más cercano en el espacio de búsqueda
            distances = np.abs(np.array(search_space) - template_scan_point)
            
            # Si la distancia mínima está dentro de la tolerancia...
            if np.min(distances) < tolerance:
                # ...encontramos su posición
                best_local_idx = np.argmin(distances)
                
                # Y lo asignamos
                scan_point_to_assign = search_space[best_local_idx]
                self.manual_assignments[scan_point_to_assign] = bp_size
                assigned_count += 1
                
                # IMPORTANTE: Actualizamos el índice para que la próxima búsqueda empiece DESPUÉS de este pico
                # El '+1' asegura que no volvemos a considerar el mismo pico
                last_detected_idx += best_local_idx + 1
            
        if not silent:
            messagebox.showinfo(
                "Asignación Automática", 
                f"Se han re-asignado {assigned_count} de {len(self.first_sample_template)} picos usando la plantilla.", 
                parent=self
            )

    def save_and_next(self):
        current_full_path = self.app.fsa_files[self.current_file_index]
        if len(self.manual_assignments) < 2:
            if not messagebox.askyesno("Confirmar", "Calibración inválida (<2 puntos). ¿Quieres saltar esta muestra?", parent=self):
                return
            self.calibrations[current_full_path] = None
        else:
            num_points = len(self.manual_assignments)
            kind = 'cubic' if num_points >= 4 else ('quadratic' if num_points == 3 else 'linear')
            sorted_points = sorted(self.manual_assignments.items())
            scan_points = np.array([p[0] for p in sorted_points])
            bp_sizes = np.array([p[1] for p in sorted_points])
            if not np.all(np.diff(scan_points) > 0):
                messagebox.showerror("Error de Calibración", "Los puntos de calibración deben tener valores de escaneo crecientes.", parent=self)
                return
            calib_func = interp1d(scan_points, bp_sizes, kind=kind, fill_value="extrapolate")
            self.calibrations[current_full_path] = (calib_func, self.manual_assignments)
            if self.current_file_index == 0 and self.first_sample_template is None and len(self.manual_assignments) > 0:
                self.first_sample_template = {v: int(k) for k, v in self.manual_assignments.items()}
        self.current_file_index += 1
        if self.current_file_index < len(self.app.fsa_files):
            self.setup_for_current_file()
        else:
            self.app.finish_calibration(self.calibrations, self.first_sample_template)
            self.destroy()

    # En la clase CalibrationWizard, reemplaza esta función:

    # En la clase CalibrationWizard, reemplaza también esta función:

    def detect_peaks(self, silent=False):
        if self.raw_data is None:
            return
        try:
            ignore_until_scan = int(self.ignore_scans_var.get())
            height = float(self.height_var.get())
            prominence = float(self.prominence_var.get())
            distance = int(self.distance_var.get())
            
            if ignore_until_scan >= len(self.raw_data):
                if not silent: messagebox.showinfo("Aviso", "El valor 'Ignorar scans hasta' es mayor que la longitud de los datos.", parent=self)
                self.detected_peaks_indices = np.array([])
            else:
                data_slice = self.raw_data[ignore_until_scan:]
                indices_relative, _ = find_peaks(data_slice, height=height, prominence=prominence, distance=distance)
                self.detected_peaks_indices = indices_relative + ignore_until_scan
            
            # --- LÓGICA CORREGIDA ---
            # Si el usuario ha pulsado el botón (no es silencioso) y tenemos una plantilla...
            if not silent and self.first_sample_template:
                self.manual_assignments = {} # Limpiamos las asignaciones anteriores
                self._auto_assign_from_template(silent=False) # Re-aplicamos la plantilla

            if not silent:
                self.redraw_plot()
                
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"No se pudieron detectar los picos: {e}", parent=self)

    

    def _on_plot_click(self, event):
        """Maneja el clic del usuario en el gráfico de calibración."""
        
        # Si el clic no tiene coordenadas válidas (está fuera del gráfico) o no fue en el área de los ejes, no hacemos nada.
        if event.inaxes != self.ax or event.xdata is None:
            return

        # Encontramos el pico detectado más cercano al clic del ratón
        clicked_scan_point = event.xdata
        
        # Si no hay picos detectados, no podemos continuar
        if self.detected_peaks_indices.size == 0:
            return
            
        # Calculamos la distancia desde el punto del clic a todos los picos detectados
        distances = np.abs(self.detected_peaks_indices - clicked_scan_point)
        
        # Si el clic está demasiado lejos de cualquier pico, lo ignoramos para evitar clics accidentales.
        # El umbral (e.g., 20 puntos de escaneo) se puede ajustar si es necesario.
        if np.min(distances) > 20: 
            return

        # Obtenemos el índice y el valor del pico más cercano
        closest_peak_index_in_array = np.argmin(distances)
        scan_point_to_assign = self.detected_peaks_indices[closest_peak_index_in_array]

        # Preparamos la lista de tamaños del marcador que aún no han sido asignados
        ladder_sizes = KNOWN_LADDERS[self.app.ladder_type_var.get()]
        assigned_bps = self.manual_assignments.values()
        available_sizes = sorted(list(np.setdiff1d(ladder_sizes, list(assigned_bps))))
        
        # Añadimos la opción de borrar si el pico ya tiene una asignación
        if scan_point_to_assign in self.manual_assignments:
            available_sizes.insert(0, "BORRAR ASIGNACIÓN")

        # Mostramos un diálogo para que el usuario elija el tamaño
        dialog = SelectPeakDialog(self, "Asignar Tamaño", available_sizes) # 'self' es el padre correcto
        selected_bp = dialog.result

        # Procesamos la selección del usuario
        if selected_bp is not None:
            if selected_bp == "DELETE":
                # Si el usuario quiere borrar, eliminamos la asignación
                if scan_point_to_assign in self.manual_assignments:
                    del self.manual_assignments[scan_point_to_assign]
            else:
                # Si no, creamos o actualizamos la asignación
                self.manual_assignments[scan_point_to_assign] = selected_bp
        
        # Redibujamos el gráfico para que se vean los cambios al instante
        self.redraw_plot()
    def redraw_plot(self):
        current_xlim = self.ax.get_xlim()
        current_ylim = self.ax.get_ylim()
        is_zoomed = not self.ax.get_autoscalex_on() and not self.ax.get_autoscaley_on()
        self.ax.clear()
        if self.raw_data is not None:
            self.ax.plot(self.raw_data, color='grey', alpha=0.7, label='Datos Brutos Marcador')
            if len(self.detected_peaks_indices) > 0:
                unassigned_peaks = np.setdiff1d(self.detected_peaks_indices, list(self.manual_assignments.keys()))
                if len(unassigned_peaks) > 0:
                    self.ax.plot(unassigned_peaks, self.raw_data[unassigned_peaks], 'x', color='red', label='Picos Detectados')
            if self.manual_assignments:
                assigned_scans = np.array(list(self.manual_assignments.keys()), dtype=int)
                assigned_bps = np.array(list(self.manual_assignments.values()))
                sort_indices = np.argsort(assigned_scans)
                assigned_scans = assigned_scans[sort_indices]
                assigned_bps = assigned_bps[sort_indices]
                valid_indices = assigned_scans < len(self.raw_data)
                self.ax.plot(assigned_scans[valid_indices], self.raw_data[assigned_scans[valid_indices]], 'o', color='blue', markersize=8, fillstyle='none', markeredgewidth=1.5, label='Picos Asignados')
                for sp, bp in zip(assigned_scans[valid_indices], assigned_bps[valid_indices]):
                    self.ax.text(sp, self.raw_data[sp] + (self.raw_data.max() * 0.02), f'{bp:.0f}', color='blue', fontweight='bold', fontsize=8, ha='center')
        self.ax.set_title("Haz clic en un pico 'x' para asignarle un tamaño del marcador")
        self.ax.set_xlabel("Puntos de escaneo (Scan Points)")
        self.ax.set_ylabel("Intensidad (RFU)")
        self.ax.grid(True, linestyle=':')
        handles, labels = self.ax.get_legend_handles_labels()
        if handles:
            self.ax.legend(handles, labels)
        if is_zoomed:
            self.ax.set_xlim(current_xlim)
            self.ax.set_ylim(current_ylim)
        self.fig.tight_layout()
        self.canvas.draw()

    def clear_assignments(self, full_reset=False):
        self.manual_assignments = {}
        if full_reset:
            self.detected_peaks_indices = np.array([])
            if self.raw_data is not None:
                self.ax.autoscale()
        self.redraw_plot()

# REEMPLAZA LA CLASE ANTERIOR CON ESTA VERSIÓN MEJORADA

# REEMPLAZA LA CLASE PlotViewerWindow CON ESTA VERSIÓN FINAL

# REEMPLAZA TU CLASE PlotViewerWindow ENTERA POR ESTE BLOQUE

class PlotViewerWindow(tk.Toplevel):
    def __init__(self, master, app_instance):
        super().__init__(master)
        self.app = app_instance
        self.title("Visor de Gráficos y Tabla de Picos")
        self.geometry("1600x900")

        # --- Paneles para dividir la ventana ---
        main_pane = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)

        left_pane = ttk.PanedWindow(main_pane, orient=tk.VERTICAL)
        main_pane.add(left_pane, weight=2)

        # --- 1. Panel de Controles ---
        controls_frame = ttk.LabelFrame(left_pane, text="Controles", padding=10)
        left_pane.add(controls_frame, weight=0)

        # Controles de Canales
        self.channel_vars = {}
        all_channels_in_listbox = self.app.sample_channel_listbox.get(0, tk.END)
        initially_selected_indices = self.app.sample_channel_listbox.curselection()
        ttk.Label(controls_frame, text="Canales:").pack(side=tk.LEFT, padx=(0, 5))
        for i, channel_name in enumerate(all_channels_in_listbox):
            var = tk.BooleanVar(value=(i in initially_selected_indices))
            self.channel_vars[channel_name] = var
            display_name = CHANNEL_DISPLAY_NAME_MAP.get(channel_name, channel_name)
            cb = ttk.Checkbutton(controls_frame, text=display_name, variable=var, command=self.update_plots, style="Toolbutton")
            cb.pack(side=tk.LEFT, padx=2)
            
        ttk.Separator(controls_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, padx=10, fill='y')
        
        # Control de Superposición
        self.overlay_var = tk.BooleanVar(value=False)
        overlay_cb = ttk.Checkbutton(controls_frame, text="Superponer Muestras", variable=self.overlay_var, command=self.update_plots, style="Toolbutton")
        overlay_cb.pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(controls_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, padx=10, fill='y')

        # Controles de Detección
        ttk.Label(controls_frame, text="Altura Mínima (RFU):").pack(side=tk.LEFT, padx=(5, 0))
        self.peak_height_var = tk.StringVar(value="100")
        ttk.Entry(controls_frame, textvariable=self.peak_height_var, width=8).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(controls_frame, text="🔍 Encontrar Picos", command=self._find_and_display_peaks).pack(side=tk.LEFT, padx=5)
        ttk.Button(controls_frame, text="=> Enviar Pico a Calculadora", command=self._send_peak_to_calculator).pack(side=tk.LEFT, padx=5)
        
        # --- 2. Panel del Gráfico ---
        plot_frame = ttk.Frame(left_pane)
        left_pane.add(plot_frame, weight=1)
        self.fig = plt.Figure(figsize=(10, 8), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2Tk(self.canvas, plot_frame, pack_toolbar=False)
        self.toolbar.update()
        self.toolbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # --- 3. Panel Derecho para la Tabla de Picos ---
        table_container = ttk.Frame(main_pane)
        main_pane.add(table_container, weight=1)
        table_frame = ttk.LabelFrame(table_container, text="Tabla de Picos Detectados", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        columns = ('file', 'channel', 'size', 'height')
        self.peak_table = ttk.Treeview(table_frame, columns=columns, show='headings')
        self.peak_table.heading('file', text='Archivo'); self.peak_table.column('file', width=180)
        self.peak_table.heading('channel', text='Canal'); self.peak_table.column('channel', width=60)
        self.peak_table.heading('size', text='Tamaño (pb)'); self.peak_table.column('size', width=80)
        self.peak_table.heading('height', text='Altura (RFU)'); self.peak_table.column('height', width=80)
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.peak_table.yview)
        self.peak_table.configure(yscrollcommand=scrollbar.set)
        
        self.peak_table.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        table_frame.grid_rowconfigure(0, weight=1); table_frame.grid_columnconfigure(0, weight=1)
        
        export_button = ttk.Button(table_container, text="📊 Exportar Tabla a Excel", command=self._export_to_excel, style="Accent.TButton")
        export_button.pack(fill=tk.X, ipady=5)

        # --- Variables y final de la inicialización ---
        self.peak_markers = []
        self.last_clicked_peak = None
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.update_plots()

    def on_close(self):
        self.app.plot_viewer = None
        self.destroy()

    # En la clase PlotViewerWindow, reemplaza esta función:

    def _send_peak_to_calculator(self):
        """
        Crea o muestra la calculadora y le envía los datos del último pico seleccionado.
        La calculadora ahora será hija de esta ventana de gráficos.
        """
        if self.last_clicked_peak is None:
            messagebox.showwarning("Sin Selección", "Por favor, primero haz clic en un pico del gráfico.", parent=self)
            return
        
        # Si la calculadora no existe o se ha cerrado, la creamos de nuevo.
        # CRUCIAL: Pasamos 'self' (la ventana de gráficos) como su nuevo padre.
        if self.app.calculator is None or not self.app.calculator.winfo_exists():
            self.app.calculator = FormulaCalculator(self)
        
        # Le enviamos los datos del pico y la traemos al frente
        self.app.calculator.add_peak(self.last_clicked_peak)
        self.app.calculator.lift()
    # En la clase PlotViewerWindow, reemplaza la función _find_and_display_peaks por esta:

    def _find_and_display_peaks(self):
        # La llamada a update_plots ya no necesita el argumento clear_peak_markers
        self.update_plots(clear_table=True) # <--- LÍNEA CORREGIDA
        
        # El resto de la función se queda igual
        try:
            min_height = float(self.peak_height_var.get())
        except ValueError:
            messagebox.showerror("Error", "La altura mínima del pico debe ser un número.", parent=self)
            return
            
        selected_sample_channels = [name for name, var in self.channel_vars.items() if var.get()]
        selected_file_indices = self.app.file_listbox.curselection()
        files_to_plot = [self.app.fsa_files[i] for i in selected_file_indices]
        
        axes_to_process = self.fig.axes
        is_overlay = self.overlay_var.get()
        
        for ax_idx, ax in enumerate(axes_to_process):
            files_for_this_ax = files_to_plot if is_overlay else [files_to_plot[ax_idx]]
            
            for full_path in files_for_this_ax:
                filename_key = Path(full_path).name
                calib_data = self.app.calibrations.get(full_path)
                if not calib_data: continue
                
                local_calib_func, _ = calib_data
                
                for channel_name in selected_sample_channels:
                    if channel_name in self.app.loaded_data.get(filename_key, {}):
                        y_raw = self.app.loaded_data[filename_key][channel_name]
                        y_cleaned = self._clean_trace_hammock(y_raw)
                        
                        indices, props = find_peaks(y_cleaned, height=min_height, prominence=min_height/4)
                        
                        if len(indices) > 0:
                            sizes_bp = local_calib_func(indices)
                            heights_rfu = props['peak_heights']
                            
                            channel_display_name = CHANNEL_DISPLAY_NAME_MAP.get(channel_name, channel_name)
                            for size, height in zip(sizes_bp, heights_rfu):
                                table_values = (filename_key, channel_display_name, f"{size:.1f}", f"{height:.0f}")
                                self.peak_table.insert('', tk.END, values=table_values)
                            
                            color = CHANNEL_COLOR_MAP.get(channel_name, 'purple')
                            marker = ax.plot(sizes_bp, heights_rfu, 'v', markersize=5, alpha=0.7, color=color)[0]
                            self.peak_markers.append(marker)
    
        self.canvas.draw()
    
    def _export_to_excel(self):
        if not self.peak_table.get_children():
            messagebox.showwarning("Exportar", "No hay picos en la tabla para exportar.", parent=self); return
        filepath = filedialog.asksaveasfilename(title="Guardar como Excel",defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
        if not filepath: return
        try:
            workbook = openpyxl.Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "Resumen de Picos"
            headers = [self.peak_table.heading(c)['text'] for c in self.peak_table['columns']]
            summary_sheet.append(headers)
            all_peaks = []
            for item_id in self.peak_table.get_children():
                values = self.peak_table.item(item_id)['values']
                summary_sheet.append(values)
                all_peaks.append(dict(zip(headers, values)))
            pivoted_data = {}
            for peak in all_peaks:
                filename = peak['Archivo']
                channel = peak['Canal']
                size = peak['Tamaño (pb)']
                if filename not in pivoted_data: pivoted_data[filename] = {}
                if channel not in pivoted_data[filename]: pivoted_data[filename][channel] = []
                pivoted_data[filename][channel].append(float(size))
            for filename, channel_data in pivoted_data.items():
                sheet_name = filename.split('.')[0][:30]
                sample_sheet = workbook.create_sheet(title=sheet_name)
                column_headers = list(channel_data.keys())
                column_data = list(channel_data.values())
                sample_sheet.append(column_headers)
                for row_data in zip_longest(*column_data, fillvalue=""):
                    sample_sheet.append(row_data)
            params_sheet = workbook.create_sheet(title="Parámetros de Análisis")
            params = {
                "Archivos Analizados": ", ".join([Path(f).name for i, f in enumerate(self.app.fsa_files) if i in self.app.file_listbox.curselection()]),
                "Tipo de Marcador": self.app.ladder_type_var.get(), "Canal del Marcador": self.app.ladder_channel_var.get(),
                "Altura Mínima (RFU) para Detección": self.peak_height_var.get(),
                "Fecha de Análisis": np.datetime_as_string(np.datetime64('now', 's'), unit='s')
            }
            params_sheet.append(["Parámetro", "Valor"])
            for key, value in params.items():
                params_sheet.append([key, value])
            workbook.save(filepath)
            messagebox.showinfo("Éxito", f"Resultados exportados a:\n{filepath}", parent=self)
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo guardar el archivo de Excel:\n{e}", parent=self)
            
    @staticmethod
    def _clean_trace_hammock(y_data, num_chunks=30):
        if len(y_data) < num_chunks * 2: return y_data
        x_points = np.arange(len(y_data)); chunk_size = len(y_data) // num_chunks
        if chunk_size == 0: return y_data
        anchor_x = [0]; anchor_y = [y_data[0]]
        for i in range(num_chunks):
            start = i * chunk_size; end = start + chunk_size
            chunk = y_data[start:end]
            if chunk.size > 0:
                min_index_in_chunk = np.argmin(chunk)
                anchor_x.append(start + min_index_in_chunk); anchor_y.append(chunk[min_index_in_chunk])
        anchor_x.append(len(y_data) - 1); anchor_y.append(y_data[-1])
        baseline = np.interp(x_points, anchor_x, anchor_y)
        cleaned_data = y_data - baseline; cleaned_data[cleaned_data < 0] = 0
        return cleaned_data

    # En la clase PlotViewerWindow, reemplaza esta función:

    def update_plots(self, clear_table=False):
        self.fig.clear()
        
        # --- CORRECCIÓN ---
        # Borramos las referencias a los marcadores antiguos. 
        # Los gráficos ya se han borrado con self.fig.clear().
        self.peak_markers = []
        
        if clear_table:
            self.peak_table.delete(*self.peak_table.get_children())
        
        # El resto de la función se queda exactamente igual
        try:
            selected_sample_channels = [name for name, var in self.channel_vars.items() if var.get()]
            selected_file_indices = self.app.file_listbox.curselection()
            if not selected_file_indices:
                self.fig.text(0.5, 0.5, "No hay muestras seleccionadas.", ha='center'); self.canvas.draw(); return
            files_to_plot = [self.app.fsa_files[i] for i in selected_file_indices]
            xlim_min = float(self.app.xlim_min_var.get()); xlim_max = float(self.app.xlim_max_var.get())
        except (ValueError, TypeError):
            messagebox.showerror("Error", "Parámetros de visualización inválidos.", parent=self); return
        if not selected_sample_channels:
            self.fig.text(0.5, 0.5, "Ningún canal seleccionado.", ha='center'); self.canvas.draw(); return

        self.app._clear_annotations()
        self.fig.canvas.mpl_connect('pick_event', self.app._on_plot_click)
        ladder_channel = self.app.ladder_channel_var.get()
        
        is_overlay = self.overlay_var.get()
        if is_overlay:
            # --- MODO SUPERPOSICIÓN: 1 GRÁFICO ---
            ax = self.fig.add_subplot(111)
            ax.set_title("Superposición de Muestras")
            for full_path in files_to_plot:
                filename_key = Path(full_path).name
                calib_data = self.app.calibrations.get(full_path)
                if not calib_data: continue
                local_calib_func, _ = calib_data
                for sample_ch in selected_sample_channels:
                    if sample_ch in self.app.loaded_data.get(filename_key, {}):
                        y_raw = self.app.loaded_data[filename_key][sample_ch]
                        y_cleaned = self._clean_trace_hammock(y_raw)
                        x_scan = np.arange(len(y_cleaned))
                        x_bp = local_calib_func(x_scan)
                        color = CHANNEL_COLOR_MAP.get(sample_ch, 'purple')
                        display_name = CHANNEL_DISPLAY_NAME_MAP.get(sample_ch, sample_ch)
                        label = f"{Path(filename_key).stem} - {display_name}"
                        line, = ax.plot(x_bp, y_cleaned, color=color, label=label, linewidth=1.2, alpha=0.7)
                        line.set_picker(5); line.full_data = (x_bp, y_cleaned)
            ax.grid(True, linestyle=':'); ax.legend(fontsize='small'); ax.set_xlim(xlim_min, xlim_max); ax.set_ylabel("RFU"); ax.set_xlabel("Tamaño (pb)")
        else:
            # --- MODO NORMAL: GRÁFICOS APILADOS ---
            num_files = len(files_to_plot)
            axs = self.fig.subplots(num_files, 1, sharex=True, squeeze=False).flatten()
            for i, full_path in enumerate(files_to_plot):
                ax = axs[i]
                filename_key = Path(full_path).name
                ax.set_title(filename_key, fontsize=10)
                calib_data = self.app.calibrations.get(full_path)
                if calib_data is None:
                    ax.text(0.5, 0.5, "Calibración saltada", color='red', ha='center', transform=ax.transAxes); continue
                local_calib_func, assigned_peaks = calib_data
                if ladder_channel in self.app.loaded_data.get(filename_key, {}):
                    y_ladder = self.app.loaded_data[filename_key][ladder_channel]
                    x_scan_ladder = np.arange(len(y_ladder)); x_bp_ladder = local_calib_func(x_scan_ladder)
                    ladder_display_name = CHANNEL_DISPLAY_NAME_MAP.get(ladder_channel, ladder_channel)
                    ax.plot(x_bp_ladder, y_ladder, color='grey', alpha=0.4, linewidth=1, label=f'Marcador ({ladder_display_name})')
                for sample_ch in selected_sample_channels:
                    if sample_ch in self.app.loaded_data.get(filename_key, {}):
                        y_sample_raw = self.app.loaded_data[filename_key][sample_ch]
                        y_sample_cleaned = self._clean_trace_hammock(y_sample_raw)
                        x_scan_sample = np.arange(len(y_sample_cleaned)); x_bp_sample = local_calib_func(x_scan_sample)
                        color = CHANNEL_COLOR_MAP.get(sample_ch, 'purple')
                        display_name = CHANNEL_DISPLAY_NAME_MAP.get(sample_ch, sample_ch)
                        line, = ax.plot(x_bp_sample, y_sample_cleaned, color=color, label=display_name, linewidth=1.2)
                        line.set_picker(5); line.full_data = (x_bp_sample, y_sample_cleaned)
                if ladder_channel in self.app.loaded_data.get(filename_key, {}):
                    y_ladder = self.app.loaded_data[filename_key][ladder_channel]
                    for scan_point, bp_size in assigned_peaks.items():
                        scan_point = int(scan_point)
                        if scan_point < len(y_ladder):
                            rfu = y_ladder[scan_point]
                            ax.vlines(x=bp_size, ymin=0, ymax=rfu, color='red', linestyle='--', alpha=0.8)
                            ax.text(bp_size, rfu, f' {int(bp_size)}', color='red', fontsize=8, ha='center', va='bottom')
                ax.grid(True, linestyle=':'); ax.legend(fontsize='small'); ax.set_xlim(xlim_min, xlim_max); ax.set_ylabel("RFU")
            if num_files > 0: axs[-1].set_xlabel("Tamaño (pb)")
        
        self.fig.suptitle("Análisis de Fragmentos Multicanal", fontsize=16)
        self.fig.tight_layout(rect=[0, 0, 1, 0.96])
        self.canvas.draw()
# --- Ventana de la Calculadora de Fórmulas ---
# --- Ventana de la Calculadora de Fórmulas ---
class FormulaCalculator(tk.Toplevel):
    # En la clase FormulaCalculator, REEMPLAZA esta función __init__ completa

    def __init__(self, master):
        super().__init__(master)
        self.title("Calculadora de Fórmulas")
        self.geometry("550x400")
        self.transient(master)
        self.grab_set()

        self.variables = {}

        # --- Creación de los frames (sin empaquetarlos aún) ---
        vars_frame = ttk.LabelFrame(self, text="Variables Asignadas", padding=10)
        calc_frame = ttk.LabelFrame(self, text="Cálculo", padding=10)
        button_frame = ttk.Frame(self, padding=(10, 10))

        # --- LÓGICA DE EMPAQUETADO CORREGIDA ---
        # 1. Empaquetamos los botones en la parte INFERIOR primero.
        button_frame.pack(side=tk.BOTTOM, fill=tk.X)
        # 2. Empaquetamos la fórmula justo ENCIMA de los botones.
        calc_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)
        # 3. Empaquetamos la tabla para que ocupe TODO el espacio restante.
        vars_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # --- Contenido de la tabla de variables ---
        columns = ('var', 'rfu', 'size', 'sample')
        self.vars_table = ttk.Treeview(vars_frame, columns=columns, show='headings')
        self.vars_table.heading('var', text='Variable'); self.vars_table.column('var', width=60, anchor='center')
        self.vars_table.heading('rfu', text='RFU'); self.vars_table.column('rfu', width=80, anchor='center')
        self.vars_table.heading('size', text='Tamaño (pb)'); self.vars_table.column('size', width=100, anchor='center')
        self.vars_table.heading('sample', text='Muestra'); self.vars_table.column('sample', width=200)
        
        # Añadimos la scrollbar a la tabla
        scrollbar = ttk.Scrollbar(vars_frame, orient=tk.VERTICAL, command=self.vars_table.yview)
        self.vars_table.configure(yscrollcommand=scrollbar.set)
        
        self.vars_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Contenido del frame de la fórmula ---
        ttk.Label(calc_frame, text="Fórmula:").grid(row=0, column=0, padx=5, sticky='w')
        self.formula_var = tk.StringVar(value="A / (A + B) * 100")
        self.formula_entry = ttk.Entry(calc_frame, textvariable=self.formula_var, font=("Segoe UI", 10))
        self.formula_entry.grid(row=0, column=1, padx=5, sticky='ew')
        self.result_label = ttk.Label(calc_frame, text="Resultado: -", font=("Segoe UI", 12, "bold"))
        self.result_label.grid(row=1, column=0, columnspan=2, pady=10, sticky='w', padx=5)
        calc_frame.columnconfigure(1, weight=1)

        # --- Contenido del frame de botones ---
        ttk.Button(button_frame, text="Limpiar Todo", command=self.clear_all).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Calcular", command=self._calculate, style="Accent.TButton").pack(side=tk.RIGHT)
        ttk.Button(button_frame, text="Cerrar", command=self.destroy).pack(side=tk.RIGHT, padx=5)

    def add_peak(self, peak_data):
        """Añade un pico a la lista de variables de la calculadora."""
        self.lift()
        variable_name = simpledialog.askstring("Asignar Variable", "Introduce un nombre para esta variable (ej: A, B, Donante):", parent=self)
        
        if not variable_name or not variable_name.strip():
            return
            
        variable_name = variable_name.strip().replace(" ", "_")
        self.variables[variable_name] = peak_data
        
        self.update_table()
        self.lift()

    def update_table(self):
        """Limpia y rellena la tabla con las variables actuales."""
        self.vars_table.delete(*self.vars_table.get_children())
        for name, data in self.variables.items():
            values = (name, f"{data['rfu']:.0f}", f"{data['size']:.1f}", f"{data['sample']} ({data['channel']})")
            self.vars_table.insert('', tk.END, values=values)

    def clear_all(self):
        """Limpia todas las variables y la tabla."""
        self.variables = {}
        self.result_label.config(text="Resultado: -")
        self.update_table()

    def _calculate(self):
        """Evalúa la fórmula introducida por el usuario."""
        formula = self.formula_var.get()
        
        try:
            temp_formula = formula
            for name, data in self.variables.items():
                import re
                temp_formula = re.sub(r'\b' + re.escape(name) + r'\b', str(data['rfu']), temp_formula)

            # Medida de seguridad básica
            allowed_chars = "0123456789.+-*/() "
            if not all(char in allowed_chars for char in temp_formula):
                raise ValueError("Caracteres no permitidos")

            result = eval(temp_formula)
            self.result_label.config(text=f"Resultado: {result:.4f}")

        except Exception as e:
            self.result_label.config(text=f"Error: {e}")
# --- Clase Principal ---
class AnalizadorFSA:
    def __init__(self, master):
        self.master = master
        self.master.title("Visor de Fragmentos FSA v11.0")
        self.fsa_files = []
        self.loaded_data = {}
        self.calibrations = {}
        self.calibration_template = None
        self.setup_styles()
        self.create_widgets()
        self._create_menubar() # <-- AÑADE ESTA LÍNEA AL FINAL
        self.plot_annotations = {} # <-- AÑADE ESTA LÍNEA
        self.plot_viewer = None
        self.calculator = None # <-- AÑADE ESTA LÍNEA 

    # En la clase AnalizadorFSA, reemplaza esta función:

    # En la clase AnalizadorFSA
    
    def setup_styles(self):
        self.style = ttk.Style()
        self.title_font = ("Segoe UI", 12, "bold")
        self.label_font = ("Segoe UI", 10)
        self.status_font = ("Segoe UI", 9, "italic")
        self.style.configure("TLabelFrame", padding=10)
        self.style.configure("TLabelFrame.Label", font=self.title_font, foreground="#333")
        self.style.configure("Accent.TButton", font=self.label_font, padding=5)
        
        # La línea para "Enter.TButton" se ha eliminado de aquí porque ya no es necesaria.

    # En la clase AnalizadorFSA, reemplaza esta función:

    def create_widgets(self):
        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- SECCIÓN 1: Carga de Archivos ---
        load_frame = ttk.LabelFrame(main_frame, text="1. Cargar Archivos")
        load_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(load_frame, text="Seleccionar Archivos .fsa", command=self.select_fsa_files, style="Accent.TButton").pack(fill=tk.X, ipady=4)
        self.fsa_file_label = ttk.Label(load_frame, text="0 archivos seleccionados", font=self.label_font)
        self.fsa_file_label.pack(pady=5)

        # --- SECCIÓN 2: Calibración ---
        calib_frame = ttk.LabelFrame(main_frame, text="2. Calibración")
        calib_frame.pack(fill=tk.X, padx=5, pady=5, ipady=5)
        # (El contenido de este frame no cambia, lo omito por brevedad, pero debe estar aquí)
        # ... (Aquí va todo el código para los menús de canal, tipo de marcador, plantillas, etc.) ...
        calib_grid = ttk.Frame(calib_frame, padding=5); calib_grid.pack(fill=tk.X)
        ttk.Label(calib_grid, text="Canal del Marcador:", font=self.label_font).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.ladder_channel_var = tk.StringVar()
        self.ladder_channel_menu = ttk.Combobox(calib_grid, textvariable=self.ladder_channel_var, state='disabled', width=15); self.ladder_channel_menu.grid(row=0, column=1, padx=5, sticky="ew")
        ttk.Label(calib_grid, text="Tipo de Marcador:", font=self.label_font).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.ladder_type_var = tk.StringVar(value=list(KNOWN_LADDERS.keys())[0])
        ttk.Combobox(calib_grid, textvariable=self.ladder_type_var, values=list(KNOWN_LADDERS.keys()), state='readonly', width=25).grid(row=1, column=1, padx=5, sticky="ew")
        calib_grid.columnconfigure(1, weight=1)
        ttk.Separator(calib_frame, orient='horizontal').pack(fill='x', pady=10, padx=5)
        template_frame = ttk.Frame(calib_frame, padding=5); template_frame.pack(fill=tk.X)
        self.load_template_button = ttk.Button(template_frame, text="Cargar Plantilla", command=self.load_template, state='disabled'); self.load_template_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        self.save_template_button = ttk.Button(template_frame, text="Guardar Plantilla", command=self.save_template, state='disabled'); self.save_template_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(5, 0))
        self.template_status_label = ttk.Label(calib_frame, text="Plantilla: No cargada", font=self.status_font); self.template_status_label.pack(pady=5)
        self.calibrate_button = ttk.Button(calib_frame, text="Iniciar Proceso de Calibración", command=self.start_calibration_process, state='disabled', style="Accent.TButton"); self.calibrate_button.pack(fill=tk.X, pady=5, padx=5, ipady=4)
        self.calibration_status_label = ttk.Label(calib_frame, text="Estado: Pendiente de calibración", foreground="red", font=self.status_font); self.calibration_status_label.pack()

        # --- SECCIÓN 3: Selección de Muestras a Visualizar (NUEVO) ---
        selection_frame = ttk.LabelFrame(main_frame, text="3. Selección de Muestras a Visualizar")
        selection_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=10)
        
        self.file_listbox = tk.Listbox(selection_frame, selectmode=tk.EXTENDED, height=6, font=self.label_font)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=5, padx=5)

        listbox_buttons_frame = ttk.Frame(selection_frame)
        listbox_buttons_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,5))
        ttk.Button(listbox_buttons_frame, text="Todas", command=lambda: self.file_listbox.select_set(0, tk.END)).pack(pady=2)
        ttk.Button(listbox_buttons_frame, text="Ninguna", command=lambda: self.file_listbox.select_clear(0, tk.END)).pack(pady=2)


        # --- SECCIÓN 4: Opciones de Gráfico ---
        plot_options_frame = ttk.LabelFrame(main_frame, text="4. Opciones de Gráfico")
        plot_options_frame.pack(fill=tk.X, padx=5, pady=5)
        plot_grid = ttk.Frame(plot_options_frame, padding=5); plot_grid.pack(fill=tk.X)
        # (Aquí va el contenido que antes estaba en "Visualización y Limpieza")
        # ... (Canales de muestra, rango X, etc.) ...
        ttk.Label(plot_grid, text="Canales de Muestra:", font=self.label_font).grid(row=0, column=0, sticky="nw", padx=5, pady=2)
        listbox_frame = ttk.Frame(plot_grid); listbox_frame.grid(row=0, column=1, sticky="ew", rowspan=2, padx=5)
        self.sample_channel_listbox = tk.Listbox(listbox_frame, selectmode=tk.EXTENDED, height=3, exportselection=False, font=self.label_font); self.sample_channel_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(plot_grid, text="Rango Eje X (pb):", font=self.label_font).grid(row=0, column=2, sticky="w", padx=(20, 5), pady=5)
        xlim_frame = ttk.Frame(plot_grid); xlim_frame.grid(row=0, column=3, sticky="w", padx=5)
        self.xlim_min_var = tk.StringVar(value="50"); ttk.Entry(xlim_frame, textvariable=self.xlim_min_var, width=7).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(xlim_frame, text="a").pack(side=tk.LEFT, padx=5)
        self.xlim_max_var = tk.StringVar(value="500"); ttk.Entry(xlim_frame, textvariable=self.xlim_max_var, width=7).pack(side=tk.LEFT, padx=5)
        
        # Botón final
        self.visualize_button = ttk.Button(main_frame, text="Generar Gráficos para Muestras Seleccionadas", command=self.trigger_plot_generation, state='disabled', style="Accent.TButton")
        self.visualize_button.pack(fill=tk.X, padx=5, pady=15, ipady=8)

    def _create_menubar(self):
        """Crea la barra de menú superior de la aplicación."""
        menubar = tk.Menu(self.master)
        self.master.config(menu=menubar)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=file_menu)
        
        file_menu.add_command(label="Guardar Sesión...", command=self._save_session)
        file_menu.add_command(label="Cargar Sesión...", command=self._load_session)
        file_menu.add_separator()
        file_menu.add_command(label="Salir", command=self.master.quit)
        

   
    # Pega este bloque completo dentro de la clase AnalizadorFSA

    def _clear_annotations(self):
        """Borra todas las etiquetas de picos de los gráficos anteriores."""
        for subplot_ax, annotation_list in self.plot_annotations.items():
            for annotation in annotation_list:
                try:
                    annotation.remove()
                except:
                    # El objeto puede haber sido eliminado ya, lo ignoramos
                    pass
        self.plot_annotations = {}

    # En la clase AnalizadorFSA, reemplaza esta función:

    # En la clase AnalizadorFSA, reemplaza esta función:

    def _on_plot_click(self, event):
        """Se activa cuando el usuario hace clic en una línea de muestra en el gráfico final."""
        artist = event.artist
        # Ahora buscamos 'full_data', que contiene la señal completa de la línea
        if not hasattr(artist, 'full_data'):
            return

        x_data, y_data = artist.full_data
        ax = artist.axes
        
        # Limpiamos la etiqueta anterior de este subplot
        if ax in self.plot_annotations:
            for annotation in self.plot_annotations[ax]:
                annotation.remove()
        
        # --- Lógica de búsqueda de pico (sin cambios) ---
        mouse_x = event.mouseevent.xdata
        if mouse_x is None: return # Evita errores si el clic es fuera del área de datos
        
        closest_index = np.argmin(np.abs(x_data - mouse_x))
        search_radius = 20
        start = max(0, closest_index - search_radius)
        end = min(len(y_data), closest_index + search_radius)
        
        peak_index_in_slice = np.argmax(y_data[start:end])
        true_peak_index = start + peak_index_in_slice

        x_coord = x_data[true_peak_index]
        y_coord = y_data[true_peak_index]

        # --- Creación de la etiqueta (sin cambios) ---
        label_text = f"{x_coord:.1f} pb\n{y_coord:.0f} RFU"
        
        annotation = ax.annotate(label_text,
                                  xy=(x_coord, y_coord),
                                  xytext=(0, 25),
                                  textcoords="offset points",
                                  ha='center', va='bottom',
                                  bbox=dict(boxstyle="round,pad=0.5", fc="yellow", alpha=0.9),
                                  arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0"))

        self.plot_annotations[ax] = [annotation]

        # --- AÑADE ESTE BLOQUE AL FINAL ---
        # Guardamos los datos del pico clicado para poder enviarlos a la calculadora.
        # Nos aseguramos de que la ventana del gráfico exista antes de intentar acceder a ella.
        if self.plot_viewer and self.plot_viewer.winfo_exists():
            peak_info = {
                'rfu': y_coord,
                'size': x_coord,
                'sample': ax.get_title(),  # El título del subplot es el nombre del archivo
                'channel': artist.get_label() # La etiqueta de la línea es el nombre del canal (ej: 'Azul')
            }
            # Guardamos la información en la variable de la ventana del gráfico
            self.plot_viewer.last_clicked_peak = peak_info
        # --- FIN DEL BLOQUE AÑADIDO ---
        
        # Actualizamos el lienzo para que se vea la nueva etiqueta
        artist.get_figure().canvas.draw_idle()

    def _save_session(self):
        """Guarda el estado actual de la calibración en un archivo."""
        if not self.calibrations:
            messagebox.showwarning("Guardar Sesión", "No hay ninguna calibración que guardar.", parent=self.master)
            return

        filepath = filedialog.asksaveasfilename(
            title="Guardar Sesión de Análisis",
            defaultextension=".pkl",
            filetypes=[("Archivos de Sesión de Análisis", "*.pkl")]
        )
        if not filepath:
            return

        # Preparamos los datos que queremos guardar
        session_data = {
            "fsa_files": self.fsa_files,
            "calibrations": self.calibrations,
            "ladder_channel": self.ladder_channel_var.get(),
            "ladder_type": self.ladder_type_var.get()
        }

        try:
            with open(filepath, 'wb') as f: # 'wb' es importante: modo de escritura binaria
                pickle.dump(session_data, f)
            messagebox.showinfo("Guardar Sesión", "La sesión se ha guardado correctamente.", parent=self.master)
        except Exception as e:
            messagebox.showerror("Error al Guardar", f"No se pudo guardar la sesión:\n{e}", parent=self.master)

    # En tu clase AnalizadorFSA, reemplaza esta función completa:

    # En tu clase AnalizadorFSA, reemplaza esta función completa:

    def _load_session(self):
        """Carga un estado de calibración desde un archivo."""
        filepath = filedialog.askopenfilename(
            title="Cargar Sesión de Análisis",
            filetypes=[("Archivos de Sesión de Análisis", "*.pkl")]
        )
        if not filepath:
            return

        try:
            with open(filepath, 'rb') as f:
                session_data = pickle.load(f)
            
            # 1. Restaurar el estado del programa
            self.fsa_files = session_data.get("fsa_files", [])
            self.calibrations = session_data.get("calibrations", {})
            self.ladder_channel_var.set(session_data.get("ladder_channel", ""))
            self.ladder_type_var.set(session_data.get("ladder_type", list(KNOWN_LADDERS.keys())[0]))

            # 2. Limpiar y repoblar la lista visual de archivos
            self.file_listbox.delete(0, tk.END)
            
            if self.fsa_files:
                print("--- Iniciando carga de archivos en la lista ---") # Mensaje de inicio
                
                # Bucle para añadir cada archivo a la lista
                for fsa_file_path in self.fsa_files:
                    filename = Path(fsa_file_path).name
                    
                    # ----- LÍNEA DE DEPURACIÓN CLAVE -----
                    # Esto imprimirá cada nombre de archivo en la consola o terminal desde donde ejecutas el script.
                    print(f"Añadiendo a la lista: {filename}")
                    
                    self.file_listbox.insert(tk.END, filename)

                print(f"--- Carga finalizada. Total de archivos procesados: {len(self.fsa_files)} ---")

                self.file_listbox.select_set(0, tk.END)
            
            # 3. Forzar actualización de la interfaz
            # Esta línea puede resolver problemas donde la UI no se refresca correctamente.
            self.master.update_idletasks()

            # 4. Procesar datos y actualizar el resto de la UI
            self.fsa_file_label.config(text=f"{len(self.fsa_files)} archivos cargados desde sesión")
            self.process_files()
            self.update_ui_state()

            calibrated_count = sum(1 for cal in self.calibrations.values() if cal is not None)
            self.calibration_status_label.config(text=f"Estado: {calibrated_count} calibraciones cargadas", foreground="green")
            
            messagebox.showinfo("Cargar Sesión", "La sesión se ha cargado correctamente.\nYa puedes generar los gráficos.", parent=self.master)

        except Exception as e:
            messagebox.showerror("Error al Cargar", f"No se pudo cargar la sesión:\n{e}", parent=self.master)

    def load_template(self):
        filepath = filedialog.askopenfilename(title="Cargar Plantilla de Calibración", filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
        if not filepath: return
        try:
            with open(filepath, 'r') as f: data = json.load(f)
            self.calibration_template = {float(k): int(v) for k, v in data.items()}
            self.template_status_label.config(text=f"Plantilla: {Path(filepath).name}", foreground="blue")
            messagebox.showinfo("Éxito", f"Plantilla '{Path(filepath).name}' cargada correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar la plantilla:\n{e}")
            self.calibration_template = None; self.template_status_label.config(text="Plantilla: Error de carga", foreground="red")
        self.update_ui_state()

    def save_template(self):
        if not self.calibration_template: messagebox.showwarning("Aviso", "No hay ninguna plantilla para guardar."); return
        filepath = filedialog.asksaveasfilename(title="Guardar Plantilla de Calibración", defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if not filepath: return
        try:
            template_limpio = {str(k): int(v) for k, v in self.calibration_template.items()}
            with open(filepath, 'w') as f: json.dump(template_limpio, f, indent=4)
            messagebox.showinfo("Éxito", f"Plantilla guardada correctamente en '{Path(filepath).name}'.")
        except Exception as e: messagebox.showerror("Error", f"No se pudo guardar la plantilla:\n{e}")

    def start_calibration_process(self):
        if not self.fsa_files: messagebox.showerror("Error", "No hay archivos cargados."); return
        if not self.ladder_channel_var.get(): messagebox.showerror("Error", "Debes seleccionar un canal de marcador válido."); return
        CalibrationWizard(self.master, self, template=self.calibration_template)

    def finish_calibration(self, calibrations, new_template):
        self.calibrations = calibrations
        if new_template:
            self.calibration_template = new_template
            self.template_status_label.config(text="Plantilla: Lista para guardar", foreground="green")
        calibrated_count = sum(1 for cal in calibrations.values() if cal is not None)
        status_text = f"Estado: Calibradas {calibrated_count} de {len(self.fsa_files)} muestras."
        color = "green" if calibrated_count == len(self.fsa_files) else "orange"
        self.calibration_status_label.config(text=status_text, foreground=color)
        self.update_ui_state()
        
    def select_fsa_files(self):
        self.fsa_files = filedialog.askopenfilenames(title="Selecciona archivos .fsa", filetypes=[("FSA files", "*.fsa")])
        # --- AÑADIR ESTE BLOQUE ---
        self.file_listbox.delete(0, tk.END) # Limpiamos la lista por si había algo antes
        if self.fsa_files:
            for fsa_file_path in self.fsa_files:
                # Mostramos solo el nombre del archivo, no toda la ruta
                self.file_listbox.insert(tk.END, Path(fsa_file_path).name)
            self.file_listbox.select_set(0, tk.END) # Por defecto, seleccionamos todos
        # --- FIN DEL BLOQUE A AÑADIR ---
        if not self.fsa_files:
            self.fsa_file_label.config(text="0 archivos seleccionados"); self.loaded_data = {}; self.calibrations = {}; self.calibration_template = None
        else:
            self.fsa_file_label.config(text=f"{len(self.fsa_files)} archivos seleccionados"); self.process_files()
        self.update_ui_state()

    def update_ui_state(self):
        files_loaded = bool(self.fsa_files)
        calibrated_at_least_one = any(cal is not None for cal in self.calibrations.values())
        self.calibrate_button.config(state="normal" if files_loaded else "disabled")
        self.ladder_channel_menu.config(state='readonly' if files_loaded else "disabled")
        self.load_template_button.config(state="normal" if files_loaded else "disabled")
        self.save_template_button.config(state="normal" if self.calibration_template and calibrated_at_least_one else "disabled")
        self.visualize_button.config(state="normal" if calibrated_at_least_one else "disabled")
        if not files_loaded:
            self.template_status_label.config(text="Plantilla: No cargada", foreground="black")
            self.calibration_status_label.config(text="Estado: Pendiente de calibración", foreground="red")

    # Reemplaza esta función en la clase AnalizadorFSA
    def trigger_plot_generation(self):
        """
        Abre la ventana del visor de gráficos si no existe, o la actualiza si ya está abierta.
        """
        # Si la ventana no existe o ha sido cerrada
        if self.plot_viewer is None or not self.plot_viewer.winfo_exists():
            # Creamos una nueva instancia de nuestra ventana de gráficos
            self.plot_viewer = PlotViewerWindow(self.master, self)
        else:
            # Si ya existe, simplemente la traemos al frente y la refrescamos
            self.plot_viewer.update_plots()
            self.plot_viewer.lift()

    
    def process_files(self):
        self.loaded_data = {}; all_channels = set(); failed_files = []
        for f in self.fsa_files:
            try:
                with open(f, 'rb') as handle: record = SeqIO.read(handle, 'abi')
                filename = Path(f).name; self.loaded_data[filename] = {}
                abif_raw = record.annotations.get('abif_raw', {})
                available_channels = [key for key in abif_raw.keys() if key.startswith('DATA')]
                all_channels.update(available_channels)
                for channel in available_channels: self.loaded_data[filename][channel] = np.array(abif_raw[channel])
            except Exception as e: failed_files.append(f"{Path(f).name}: {e}")
        if failed_files: messagebox.showwarning("Error de Archivo", "No se pudieron procesar algunos archivos:\n\n" + "\n".join(failed_files))
        
        canales_muestra_deseados = ['DATA9', 'DATA10', 'DATA11']
        canales_marcador_deseados = ['DATA4', 'DATA105']
        self.sample_channel_listbox.delete(0, tk.END)
        for channel in canales_muestra_deseados:
             if channel in all_channels: self.sample_channel_listbox.insert(tk.END, channel)
        
        ladder_channels_filtrados = [ch for ch in canales_marcador_deseados if ch in all_channels]
        if not ladder_channels_filtrados:
            ladder_channels_filtrados = sorted(list(all_channels), key=lambda x: int(x[4:]))
        
        self.ladder_channel_menu['values'] = ladder_channels_filtrados
        if 'DATA4' in ladder_channels_filtrados: self.ladder_channel_var.set('DATA4')
        elif ladder_channels_filtrados: self.ladder_channel_var.set(ladder_channels_filtrados[0])

if __name__ == "__main__":
    # La magia empieza aquí: usamos ThemedTk si está disponible
    try:
        from ttkthemes import ThemedTk
        # Creamos la ventana raíz
        root = ThemedTk(theme="arc", toplevel=True, themebg=True)
    except ImportError:
        root = tk.Tk()

    # --- CORRECCIÓN: Definir el estilo ANTES de crear cualquier ventana ---
    # Creamos el objeto de estilo y configuramos el estilo de los botones de la portada.
    style = ttk.Style(root)
    style.configure("Enter.TButton", font=("Segoe UI", 16, "bold"), padding=14)
    # --- FIN DE LA CORRECCIÓN ---

    # 1. Ocultar la ventana principal inmediatamente
    root.withdraw() 
    
    # 2. Definimos la función que lanzará la aplicación principal
    def launch_main_app():
        splash.destroy()
        root.deiconify()
        
        try:
            path_icono = resource_path("icono.png")
            root.iconphoto(True, tk.PhotoImage(file=path_icono))
        except tk.TclError:
            print("No se encontró 'icono.png', se usará el icono por defecto.")
        
        # ¡IMPORTANTE! La instancia de la app se sigue creando aquí
        app = AnalizadorFSA(root)

    # 3. Creamos y mostramos la portada, pasándole la función de lanzamiento
    splash = SplashScreen(root, launch_callback=launch_main_app)
    
    # 4. Iniciamos el bucle principal.
    root.mainloop()